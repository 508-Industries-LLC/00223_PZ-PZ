import os
import openpyxl as xl

class TaskCodeSequenceGeneratorMain():
    def task_code(self):
        os.chdir('F:/00223-PROMFI DATABASE SYSTEM MANAGER/3.5.1 - 00223-PZ-PZ PROMFI SYSTEM')
        projects_workbook = xl.load_workbook('PROMFI 1.4.xlsx', data_only=True)
        tasks_sheet = projects_workbook['TAP.P']
        guess1 = input('guess a family task code: ')
        guess2 = input('guess a parent task code: ')
        print(f"Task Code Options are: ")

        for row in range(16, tasks_sheet.max_row + 1):
            task_code_selection = tasks_sheet.cell(row, 3)
            family_task_code = tasks_sheet.cell(row, 4)
            family_task_description = tasks_sheet.cell(row, 5)
            parent_task_code = tasks_sheet.cell(row, 6)
            parent_task_description = tasks_sheet.cell(row, 7)
            child_task_code = tasks_sheet.cell(row, 8)
            child_task_description = tasks_sheet.cell(row, 9)
            task_options_list = {
                task_code_selection.value: family_task_code.value,
                task_code_selection.value: family_task_description.value,
                task_code_selection.value: parent_task_code.value,
                task_code_selection.value: parent_task_description.value,
                task_code_selection.value: child_task_code.value,
                task_code_selection.value: child_task_description.value
            }

            # if guess1 == family_task_code.value and guess2 == parent_task_code.value:
            #     #  and guess3 == child_task_description.value
            #     print(f"{task_options_list}") # , end=" "
            #     # return task_options_list

