import os
import openpyxl as xl


# class ProjectNumber:
def project_number(): # project_family
    os.chdir('F:/00223-PROMFI DATABASE SYSTEM MANAGER/3.5.1 - 00223-PZ-PZ PROMFI SYSTEM')
    projects_workbook = xl.load_workbook('PROMFI 1.4.xlsx', data_only=True)
    projects_logger_sheet = projects_workbook['PLP.P']
    project_family_list = ['0-personal', '1-company', '2-unassigned', '3-unassigned', '5-library', '6-jobs' '4-unassigned','7-unassigned', '8-unassigned', '9-incubator']
    print(project_family_list)
    project_family = input('select a family from the list: ')
    print(f'Family selected: {project_family}')
    for row in range(4, projects_logger_sheet.max_row + 1):
        # print(row)
        projects_family_assigned = projects_logger_sheet.cell(row, 3)
        # print(projects_family_assigned.value)
        project_id_number = projects_logger_sheet.cell(row, 4)
        projects_logger_vacancies = projects_logger_sheet.cell(row, 5)
        if projects_logger_vacancies.value == ' ' and project_family == projects_family_assigned.value: # skip if cell is empty
            print(projects_family_assigned.value)
            print(f'The next vacant project id is: {project_id_number.value}')
            break


#