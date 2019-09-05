import os
import openpyxl as xl


# def task_code():
    # os.chdir('F:/00223-PROMFI DATABASE SYSTEM MANAGER/3.5.1 - 00223-PZ-PZ PROMFI SYSTEM')
projects_workbook = xl.load_workbook('Book1.xlsx', data_only=True)
tasks_sheet = projects_workbook['Sheet2']
guess1 = input('enter 1st guess: ')
guess2 = input('enter 2nd guess: ')
# for row in range(2, tasks_sheet.max_row + 1):
#     family_tasks_assigned = tasks_sheet.cell(row, 3)
#     task_code = tasks_sheet.cell(row, 4)
#     if task_family == family_tasks_assigned.value:
#         print(f'value is: {family_tasks_assigned.value}')
#         print(f'The next id is: {task_code.value}')
#         break
#
#

for i in range(2, tasks_sheet.max_row + 1):
    clz0 = tasks_sheet.cell(i, 2) # TASK CODE
    clz1 = tasks_sheet.cell(i, 4) # FAMILY TASK
    clz2 = tasks_sheet.cell(i, 6) # PARENT TASK
    clz3 = tasks_sheet.cell(i, 3) # FAMILY CODE
    clz4 = tasks_sheet.cell(i, 5) # PARENT CODE
    clz5 = tasks_sheet.cell(i, 7) # CHILD CODE
    clz6 = tasks_sheet.cell(i, 8) # CHILD TASK
    items_list = {
        clz0.value: clz1.value,
        clz0.value: clz2.value,
        clz0.value: clz6.value
    }
    # print(clz1.value)
    # print(items_list, end=" ")
    # task_family = input('enter a task family: ')
    if guess1 == clz3.value and guess2 == clz4.value:
        print(items_list, end=" ")
        # print(items_list)
        # print(f'primary: {clz1.value} & secondary: {clz2.value}')
