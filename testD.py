import os as os
import sys
import openpyxl as xl
from openpyxl.chart import BarChart, Reference

os.chdir('C:/Users/Phoen/Google Drive (508industries@gmail.com)/00085-PROGRAMMING & CODING PROJECTS/Python Programming/190821_python tests')
# wb = xl.load_workbook('Book1.xlsx', data_only=True)
wb = xl.load_workbook('Book1.xlsx')
# wb = xl.load_workbook('Book1.xlsx')
sht = wb['Sheet1']
# clz = sht['d4']
# clz = sht.cell(4,2)
cll = sht['A1']
cll.value = 'yayo'
wb.save(filename='Book1.xlsx')
# print(sht.max_row)
print(cll.value)

# for x in range(2, sht.max_row+ 1):
#     clz = sht.cell.value
    # for y in range(1, sht.max_row +1):
    #     clz = Reference(sht, max_col=y, max_row=x)
    #     print(clz.value)

# import os
# import openpyxl as xl
# os.chdir('F:/00085-PROGRAMMING & CODING PROJECTS/Python Programming/190422_python tutorial - programming with Mosh/PROMFI 2')
# projects_workbook = xl.load_workbook('Book1.xlsx', data_only=True)
# tasks_sheet = projects_workbook['Sheet2']
# task_code_selection = tasks_sheet.cell(2, 3)
# var1 = task_code_selection.value
# print(var1)