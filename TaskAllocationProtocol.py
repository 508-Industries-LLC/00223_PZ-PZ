# import sys
#
# print("Welcome to TAP (Task Allocation Protocol)")
# uin = input("Would you like to create a task ?")
# print("yes-y/n-n: ")
#
# if uin == "y":
#     print("please select an option below:")
#     print("F-Family Task, P-Parent Task, C-Child Task")
#     tktyp = input("enter your selection here: ")
#     if tktyp == "F":
#         print(f"You have selected: {tktyp}, therefore a {tktyp} task will be created")
#         print("Family Task Codes are: A-Z")
#         tkcd = input("Please select a family task code from the list above: ")
#

import sys
import os
import openpyxl as xl

class TaskCodeGen:
    def __init__(self):
    # def task_code_func(self):
        os.chdir('C:/Users/Phoen/Google Drive (508industries@gmail.com)/00223-PROMFI DATABASE SYSTEM MANAGER/3.5.1 - 00223-PZ-PZ PROMFI SYSTEM')
        projects_workbook = xl.load_workbook('PROMFI 1.4.xlsx', data_only=True)
        tasks_sheet = projects_workbook['Sheet1']
        # family_task_code = tasks_sheet.cell(4, 2)
        # return family_task_code.value
        guess1 = input('guess a family task code: ')
        guess2 = input('guess a parent task code: ')
        print(f"Task Code Options are: ")

        for row in range(4, tasks_sheet.max_row + 1):
            task_code_selection = tasks_sheet.cell(row, 2)
            family_task_code = tasks_sheet.cell(row, 3)
            family_task_description = tasks_sheet.cell(row, 4)
            parent_task_code = tasks_sheet.cell(row, 5)
            parent_task_description = tasks_sheet.cell(row, 6)
            child_task_code = tasks_sheet.cell(row, 7)
            child_task_description = tasks_sheet.cell(row, 8)
            task_options_list = {
                task_code_selection.value: family_task_code.value,
                task_code_selection.value: family_task_description.value,
                task_code_selection.value: parent_task_code.value,
                task_code_selection.value: parent_task_description.value,
                task_code_selection.value: child_task_code.value,
                task_code_selection.value: child_task_description.value
            }
            self.genie = task_options_list
        # return task_options_list
            if guess1 == family_task_code.value and guess2 == parent_task_code.value:
                #  and guess3 == child_task_description.value
                print(f"{task_options_list}") # , end=" "
                # return task_options_list


# print(TaskCodeGen().task_code_func())
